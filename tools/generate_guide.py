"""Genera le guide asta (moduli e ruoli) in xlsx, csv e md dentro `output/`.

Uso: python tools/generate_guide.py [--budget N] [--output DIR]
      [--alternative K] [--beam B] [--top T]

Legge la cache `data/` e lo stato asta (`data/asta.json`): il pool è
costituito dai giocatori rimasti, il budget è quello residuo (o `--budget`).
Nessuna rete.
"""

from __future__ import annotations

import argparse
import csv
import logging
from collections.abc import Mapping, Sequence
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from entities import GROUP_LABELS, ROLE_GROUP, Player, Role, RoleGroup, attach_stats
from fetch_fixtures import read_league_context, read_remaining_calendar
from fetch_quotazioni import read_players
from fetch_stats import read_season_stats
from guide import (
    beam_combinations,
    greedy_cover,
    k_best_rosters,
    position_candidates,
)
from projection import LeagueContext, project
from state import load_state, spent_budget, taken_urls
from utility import (
    MODULE_POSITIONS,
    MODULES,
    TeamCalendar,
    formation_positions,
    opponent_outlook,
    player_roles,
    remaining_weeks,
    team_strengths_from_players,
)

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class GuideContext:
    """Contesto condiviso per proiezioni e copertura."""

    league: LeagueContext
    calendars: Mapping[str, TeamCalendar]
    strengths: Mapping[str, float]


def _role_code(player) -> str:
    """Codici ruolo compatti (es. "E/W")."""
    return "/".join(role.value.upper() for role in player_roles(player))


def _easy_weeks(player, ctx: GuideContext) -> tuple[int, ...]:
    """Le giornate rimanenti con partita facile per il giocatore."""
    return tuple(
        opp.matchweek
        for opp in opponent_outlook(player, ctx.league, ctx.calendars, ctx.strengths)
        if opp.easy is True
    )


def _fmt_weeks(weeks: Sequence[int]) -> str:
    return ", ".join(str(week) for week in weeks) if weeks else "—"


def _load_data(budget: int | None) -> tuple[list, GuideContext, list, int]:
    """Cache + stato: (players, context, remaining, budget)."""
    players = attach_stats(read_players(), read_season_stats())
    ctx = GuideContext(
        league=read_league_context(),
        calendars=read_remaining_calendar(),
        strengths=team_strengths_from_players(players),
    )
    state = load_state()
    taken = taken_urls(state)
    remaining = [player for player in players if player.url not in taken]
    remaining_budget = state.budget - spent_budget(state) if budget is None else budget
    return players, ctx, remaining, remaining_budget


def _module_guide(
    module: str,
    remaining: Sequence,
    ctx: GuideContext,
    budget: int,
    alternatives: int,
) -> dict:
    """Le `alternatives` rose migliori per il modulo (esclusione progressiva)."""
    weeks = remaining_weeks(ctx.league, ctx.calendars)
    alternatives_list: list[dict] = []
    for squad in k_best_rosters(
        module,
        remaining,
        ctx.league,
        ctx.calendars,
        ctx.strengths,
        budget=budget,
        k=alternatives,
    ):
        xi_lines = formation_positions(module, squad.selected)
        xi = [slot.player for line in xi_lines for slot in line.positions if slot.player]
        bench = [player for player in squad.selected if player not in xi]
        alternatives_list.append(
            {
                "status": squad.status,
                "xi": xi,
                "bench": bench,
                "cost": squad.total_cost,
                "points": squad.total_points,
                "covered": squad.covered_weeks,
                "uncovered": tuple(week for week in weeks if week not in squad.covered_weeks),
            }
        )
    return {"module": module, "weeks": weeks, "alternatives": alternatives_list}


def _role_guide(
    module: str,
    remaining: Sequence,
    ctx: GuideContext,
    beam: int,
    top: int,
) -> list[dict]:
    """Per gruppo del modulo: candidati completi, greedy, combinazioni."""
    template = MODULE_POSITIONS[module]
    groups: list[dict] = []
    for group, role_names in zip(
        (RoleGroup.P, RoleGroup.D, RoleGroup.C, RoleGroup.A), template, strict=True
    ):
        group_players = [player for player in remaining if ROLE_GROUP[player.role] is group]
        positions: list[dict] = []
        candidate_lists: list[tuple[Player, ...]] = []
        for role_name in role_names:
            role = Role(role_name)
            candidates = position_candidates(
                role, group_players, ctx.league, ctx.calendars, ctx.strengths
            )
            positions.append({"role": role, "candidates": candidates})
            candidate_lists.append(candidates)
        greedy = greedy_cover(
            group_players, ctx.league, ctx.calendars, ctx.strengths, limit=len(role_names)
        )
        combinations = beam_combinations(
            candidate_lists, ctx.league, ctx.calendars, ctx.strengths, beam=beam, top=top
        )
        groups.append(
            {
                "group": group,
                "positions": positions,
                "greedy": greedy,
                "combinations": combinations,
            }
        )
    return groups


def _player_row(player, ctx: GuideContext) -> list[object]:
    """Riga tabellare comune: nome, squadra, qi, punti, giornate facili."""
    return [
        player.name,
        player.team_name,
        player.quote.qi,
        round(project(player, ctx.league).total_points, 1),
        _fmt_weeks(_easy_weeks(player, ctx)),
    ]


def _write_xlsx_moduli(guides: list[dict], ctx: GuideContext, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Confronto"
    ws.append(["Modulo", "Alternativa", "Stato", "Costo", "Coperto", "Su", "Punti", "Scoperte"])
    for guide in guides:
        for index, alt in enumerate(guide["alternatives"], start=1):
            ws.append(
                [
                    guide["module"],
                    index,
                    alt["status"],
                    alt["cost"],
                    len(alt["covered"]),
                    len(guide["weeks"]),
                    round(alt["points"], 1),
                    _fmt_weeks(alt["uncovered"]),
                ]
            )
    for guide in guides:
        ws = wb.create_sheet(guide["module"])
        for index, alt in enumerate(guide["alternatives"], start=1):
            ws.append(
                [
                    f"Alternativa {index} — costo {alt['cost']} — coperto "
                    f"{len(alt['covered'])}/{len(guide['weeks'])} — punti "
                    f"{alt['points']:.1f}"
                ]
            )
            ws.append(["Linea", "Ruolo", "Nome", "Squadra", "QI", "Punti", "Giornate facili"])
            for line in formation_positions(guide["module"], alt["xi"]):
                for slot in line.positions:
                    if slot.player is None:
                        continue
                    ws.append(
                        [
                            GROUP_LABELS[line.group],
                            slot.role.value.upper(),
                            *_player_row(slot.player, ctx),
                        ]
                    )
            if alt["bench"]:
                ws.append([])
                ws.append(
                    ["Panchina", "Ruolo", "Nome", "Squadra", "QI", "Punti", "Giornate facili"]
                )
                for player in alt["bench"]:
                    ws.append(["Panchina", _role_code(player), *_player_row(player, ctx)])
            ws.append([])
            ws.append(["Giornate scoperte:", _fmt_weeks(alt["uncovered"])])
            ws.append([])
    _autofit_workbook(wb)
    wb.save(path)


def _write_csv_moduli(guides: list[dict], ctx: GuideContext, path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "modulo",
                "alternativa",
                "tipo",
                "linea",
                "ruolo",
                "nome",
                "squadra",
                "qi",
                "punti",
                "giornate_facili",
            ]
        )
        for guide in guides:
            for index, alt in enumerate(guide["alternatives"], start=1):
                for line in formation_positions(guide["module"], alt["xi"]):
                    for slot in line.positions:
                        if slot.player is None:
                            continue
                        writer.writerow(
                            [
                                guide["module"],
                                index,
                                "titolare",
                                GROUP_LABELS[line.group],
                                slot.role.value.upper(),
                                *_player_row(slot.player, ctx),
                            ]
                        )
                for player in alt["bench"]:
                    writer.writerow(
                        [
                            guide["module"],
                            index,
                            "panchina",
                            "",
                            _role_code(player),
                            *_player_row(player, ctx),
                        ]
                    )


def _write_md_moduli(guides: list[dict], ctx: GuideContext, path: Path) -> None:
    lines = ["# Guide moduli — rose alternative per copertura giornate facili", ""]
    lines.append("| Modulo | Alternativa | Costo | Coperto | Punti | Scoperte |")
    lines.append("|---|---|---|---|---|---|")
    for guide in guides:
        for index, alt in enumerate(guide["alternatives"], start=1):
            lines.append(
                f"| {guide['module']} | {index} | {alt['cost']} | "
                f"{len(alt['covered'])}/{len(guide['weeks'])} | {alt['points']:.1f} | "
                f"{_fmt_weeks(alt['uncovered'])} |"
            )
    for guide in guides:
        lines.append("")
        lines.append(f"## Modulo {guide['module']}")
        for index, alt in enumerate(guide["alternatives"], start=1):
            lines.append("")
            lines.append(f"### Alternativa {index}")
            lines.append(
                f"Costo {alt['cost']} — coperto {len(alt['covered'])}/"
                f"{len(guide['weeks'])} — punti {alt['points']:.1f}"
            )
            lines.append("")
            lines.append("| Linea | Ruolo | Nome | Squadra | QI | Punti | Giornate facili |")
            lines.append("|---|---|---|---|---|---|---|")
            for line in formation_positions(guide["module"], alt["xi"]):
                for slot in line.positions:
                    if slot.player is None:
                        continue
                    lines.append(
                        f"| {GROUP_LABELS[line.group]} | {slot.role.value.upper()} | "
                        + " | ".join(str(value) for value in _player_row(slot.player, ctx))
                        + " |"
                    )
            if alt["bench"]:
                lines.append("")
                lines.append("**Panchina:**")
                for player in alt["bench"]:
                    lines.append(
                        f"- {_role_code(player)} — "
                        + " · ".join(str(value) for value in _player_row(player, ctx))
                    )
            lines.append("")
            lines.append(f"Giornate scoperte: {_fmt_weeks(alt['uncovered'])}")
    path.write_text("\n".join(lines), encoding="utf-8")


def _write_xlsx_ruoli(
    role_guides: list[list[dict]],
    modules: list[str],
    weeks: tuple[int, ...],
    ctx: GuideContext,
    path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Confronto"
    ws.append(["Modulo", "Gruppo", "Ruoli", "Candidati per posizione", "Combinazioni top"])
    for module, groups in zip(modules, role_guides, strict=True):
        for group in groups:
            ws.append(
                [
                    module,
                    GROUP_LABELS[group["group"]],
                    "/".join(pos["role"].value.upper() for pos in group["positions"]),
                    min(len(pos["candidates"]) for pos in group["positions"]),
                    len(group["combinations"]),
                ]
            )
    for module, groups in zip(modules, role_guides, strict=True):
        ws = wb.create_sheet(module)
        for group in groups:
            ws.append([])
            ws.append([GROUP_LABELS[group["group"]]])
            for pos in group["positions"]:
                ws.append([f"Posizione {pos['role'].value.upper()} — tutti i candidati"])
                ws.append(["Rango", "Nome", "Squadra", "QI", "Punti", "Giornate facili"])
                for rank, player in enumerate(pos["candidates"], start=1):
                    ws.append([rank, *_player_row(player, ctx)])
            ws.append([])
            ws.append(["Scelti (greedy)", "Aggiunte", "Coperte cumulative", "Costo"])
            for pick in group["greedy"]:
                ws.append(
                    [
                        pick.player.name,
                        len(pick.added_weeks),
                        len(pick.covered_weeks),
                        pick.cost,
                    ]
                )
            ws.append([])
            ws.append(["Top combinazioni", "Giocatori", "Coperti", "Punti", "Costo"])
            for index, combo in enumerate(group["combinations"], start=1):
                ws.append(
                    [
                        index,
                        ", ".join(player.name for player in combo.players),
                        len(combo.covered_weeks),
                        round(combo.points, 1),
                        combo.cost,
                    ]
                )
            ws.append([])
            ws.append(["Matrice copertura (X = partita facile)"] + [str(w) for w in weeks])
            for pick in group["greedy"]:
                easy = set(_easy_weeks(pick.player, ctx))
                ws.append([pick.player.name] + ["X" if w in easy else "" for w in weeks])
    _autofit_workbook(wb)
    wb.save(path)


def _write_csv_ruoli(
    role_guides: list[list[dict]],
    modules: list[str],
    ctx: GuideContext,
    path: Path,
) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "modulo",
                "gruppo",
                "ruolo",
                "tipo",
                "rango",
                "nome",
                "squadra",
                "qi",
                "punti",
                "giornate_facili",
                "coperti",
                "costo",
            ]
        )
        for module, groups in zip(modules, role_guides, strict=True):
            for group in groups:
                for pos in group["positions"]:
                    for rank, player in enumerate(pos["candidates"], start=1):
                        writer.writerow(
                            [
                                module,
                                GROUP_LABELS[group["group"]],
                                pos["role"].value.upper(),
                                "candidato",
                                rank,
                                *_player_row(player, ctx),
                                "",
                                "",
                            ]
                        )
                for pick in group["greedy"]:
                    writer.writerow(
                        [
                            module,
                            GROUP_LABELS[group["group"]],
                            _role_code(pick.player),
                            "scelto",
                            "",
                            *_player_row(pick.player, ctx),
                            len(pick.covered_weeks),
                            pick.cost,
                        ]
                    )
                for index, combo in enumerate(group["combinations"], start=1):
                    writer.writerow(
                        [
                            module,
                            GROUP_LABELS[group["group"]],
                            "/".join(_role_code(player) for player in combo.players),
                            "combinazione",
                            index,
                            ", ".join(player.name for player in combo.players),
                            "",
                            "",
                            round(combo.points, 1),
                            _fmt_weeks(combo.covered_weeks),
                            len(combo.covered_weeks),
                            combo.cost,
                        ]
                    )


def _write_md_ruoli(
    role_guides: list[list[dict]],
    modules: list[str],
    weeks: tuple[int, ...],
    ctx: GuideContext,
    path: Path,
) -> None:
    lines = ["# Guide ruoli — candidati e combinazioni per ruolo (moduli Mantra)", ""]
    for module, groups in zip(modules, role_guides, strict=True):
        lines.append(f"## Modulo {module}")
        for group in groups:
            lines.append("")
            lines.append(f"### {GROUP_LABELS[group['group']]}")
            for pos in group["positions"]:
                lines.append("")
                lines.append(f"**Posizione {pos['role'].value.upper()} — tutti i candidati:**")
                lines.append("| Rango | Nome | Squadra | QI | Punti | Giornate facili |")
                lines.append("|---|---|---|---|---|---|")
                for rank, player in enumerate(pos["candidates"], start=1):
                    lines.append(
                        f"| {rank} | "
                        + " | ".join(str(value) for value in _player_row(player, ctx))
                        + " |"
                    )
            lines.append("")
            lines.append("**Scelti (greedy):**")
            for pick in group["greedy"]:
                lines.append(
                    f"- {pick.player.name}: +{len(pick.added_weeks)} giornate, "
                    f"{len(pick.covered_weeks)} coperte cumulative, {pick.cost} crediti"
                )
            lines.append("")
            lines.append("**Top combinazioni per riga:**")
            lines.append("| # | Giocatori | Coperti | Punti | Costo |")
            lines.append("|---|---|---|---|---|")
            for index, combo in enumerate(group["combinations"], start=1):
                lines.append(
                    f"| {index} | {', '.join(player.name for player in combo.players)} | "
                    f"{len(combo.covered_weeks)} | {combo.points:.1f} | {combo.cost} |"
                )
            lines.append("")
            lines.append("Matrice copertura (X = partita facile):")
            lines.append("| Giocatore | " + " | ".join(str(w) for w in weeks) + " |")
            lines.append("|" + "---|" * (len(weeks) + 1))
            for pick in group["greedy"]:
                easy = set(_easy_weeks(pick.player, ctx))
                lines.append(
                    "| "
                    + pick.player.name
                    + " | "
                    + " | ".join("X" if w in easy else "" for w in weeks)
                    + " |"
                )
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def _autofit_workbook(wb: Workbook, max_width: int = 40) -> None:
    for ws in wb.worksheets:
        for column in ws.columns:
            letter = get_column_letter(column[0].column)
            width = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[letter].width = min(width + 2, max_width)


def main() -> None:
    parser = argparse.ArgumentParser(description="Genera le guide asta in output/")
    parser.add_argument(
        "--budget",
        type=int,
        default=None,
        help="Budget da usare (default: residuo dello stato asta)",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("output"),
        help="Directory di destinazione (default: output)",
    )
    parser.add_argument(
        "--alternative",
        type=int,
        default=10,
        help="Rose alternative per modulo (default: 10)",
    )
    parser.add_argument(
        "--beam",
        type=int,
        default=50,
        help="Larghezza del beam search per le combinazioni (default: 50)",
    )
    parser.add_argument(
        "--top",
        type=int,
        default=50,
        help="Combinazioni per riga da riportare (default: 50)",
    )
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(message)s")
    args.output.mkdir(parents=True, exist_ok=True)

    _, ctx, remaining, budget = _load_data(args.budget)
    logger.info(
        "Pool: %d giocatori rimasti — budget: %d crediti — alternative: %d",
        len(remaining),
        budget,
        args.alternative,
    )
    weeks = remaining_weeks(ctx.league, ctx.calendars)

    modules = list(MODULES)
    guides = [_module_guide(module, remaining, ctx, budget, args.alternative) for module in modules]
    role_guides = [_role_guide(module, remaining, ctx, args.beam, args.top) for module in modules]

    _write_xlsx_moduli(guides, ctx, args.output / "guide_moduli.xlsx")
    _write_csv_moduli(guides, ctx, args.output / "guide_moduli.csv")
    _write_md_moduli(guides, ctx, args.output / "guide_moduli.md")
    _write_xlsx_ruoli(role_guides, modules, weeks, ctx, args.output / "guide_ruoli.xlsx")
    _write_csv_ruoli(role_guides, modules, ctx, args.output / "guide_ruoli.csv")
    _write_md_ruoli(role_guides, modules, weeks, ctx, args.output / "guide_ruoli.md")
    for name in (
        "guide_moduli.xlsx",
        "guide_moduli.csv",
        "guide_moduli.md",
        "guide_ruoli.xlsx",
        "guide_ruoli.csv",
        "guide_ruoli.md",
    ):
        logger.info("Generato: %s", args.output / name)


if __name__ == "__main__":
    main()

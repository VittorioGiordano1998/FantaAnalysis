"""Report Excel della sessione (layer data, M6-T1).

Genera in memoria un workbook .xlsx con i cinque fogli previsti da
PLANNING §2 #7: rosa ottimale, giocatori rimasti, classifica per ruolo,
calendario, qualità/prezzo. `output/` è gitignored: per il download la UI
usa i bytes generati qui, `save_report` serve per il salvataggio locale.
"""

from __future__ import annotations

import logging
from collections.abc import Sequence
from datetime import datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from entities import GROUP_LABELS, ROLE_GROUP, ROLE_LABELS, Player
from optimize import SquadResult
from projection import LeagueContext, project

logger = logging.getLogger(__name__)

REPORT_DIR = Path("output")
_TOP_QP = 30
_HEADER_FONT = Font(bold=True)


def build_report(
    squad: SquadResult,
    players: Sequence[Player],
    league: LeagueContext,
    taken_urls: frozenset[str],
) -> bytes:
    """Genera il report Excel completo come bytes (M6-T1).

    Args:
        squad: rosa ottimale corrente (da `optimize_squad`).
        players: giocatori del listone con stats.
        league: contesto campionato per le proiezioni.
        taken_urls: giocatori già presi (esclusi dai fogli "rimasti").

    Returns:
        Bytes del file .xlsx.
    """
    wb = Workbook()
    wb.remove(wb.active)
    remaining = [p for p in players if p.url not in taken_urls]
    _sheet_rosa(wb, squad, league)
    _sheet_rimasti(wb, remaining, league)
    _sheet_classifica(wb, remaining, league)
    _sheet_calendario(wb, league)
    _sheet_qualita_prezzo(wb, remaining, league)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def save_report(data: bytes, path: Path | None = None) -> Path:
    """Salva il report su file (default `output/report_<timestamp>.xlsx`).

    Args:
        data: bytes prodotti da `build_report`.
        path: percorso di destinazione (default sotto `output/`).

    Returns:
        Il percorso del file scritto.
    """
    out = path or REPORT_DIR / f"report_{datetime.now():%Y%m%d_%H%M}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_bytes(data)
    logger.info("Report Excel scritto in %s", out)
    return out


def _sheet_rosa(wb: Workbook, squad: SquadResult, league: LeagueContext) -> None:
    """Foglio "Rosa ottimale": scelta, prezzo, punti attesi + totali."""
    headers = ["Nome", "Squadra", "Ruolo", "Prezzo", "Punti attesi"]
    rows = [
        [
            player.name,
            player.team_name,
            ROLE_LABELS[player.role],
            player.quote.qi or 0,
            round(project(player, league).total_points, 1),
        ]
        for player in squad.selected
    ]
    rows.append(["TOTALE", "", "", squad.total_cost, round(squad.total_points, 1)])
    _write_sheet(wb, "Rosa ottimale", headers, rows)


def _sheet_rimasti(wb: Workbook, remaining: Sequence[Player], league: LeagueContext) -> None:
    """Foglio "Rimasti": tutti i giocatori ancora disponibili con proiezioni."""
    headers = ["Nome", "Squadra", "Ruolo", "QI", "Punti/partita", "Punti stagione", "Q/P"]
    rows = [_player_row(player, league) for player in remaining]
    _write_sheet(wb, "Rimasti", headers, rows)


def _sheet_classifica(wb: Workbook, remaining: Sequence[Player], league: LeagueContext) -> None:
    """Foglio "Classifica per ruolo": raggruppati per gruppo e ruolo, per punti."""
    headers = ["Gruppo", "Ruolo", "Nome", "Squadra", "QI", "Punti stagione"]
    rows = sorted(
        (
            [
                GROUP_LABELS[ROLE_GROUP[player.role]],
                ROLE_LABELS[player.role],
                player.name,
                player.team_name,
                player.quote.qi or 0,
                round(project(player, league).total_points, 1),
            ]
            for player in remaining
        ),
        key=lambda row: (row[0], row[1], -row[5]),
    )
    _write_sheet(wb, "Classifica per ruolo", headers, rows)


def _sheet_calendario(wb: Workbook, league: LeagueContext) -> None:
    """Foglio "Calendario": prossimi 5 avversari con forza squadra."""
    headers = ["Squadra", "Avversario", "Gol fatti/partita", "Gol subiti/partita"]
    rows = []
    for team in sorted(league.teams.values(), key=lambda t: t.team_name):
        for opponent_id in team.upcoming_opponents:
            opponent = league.teams.get(opponent_id)
            if opponent is None:
                continue
            rows.append(
                [
                    team.team_name,
                    opponent.team_name,
                    _fmt_strength(opponent.gf_per_match),
                    _fmt_strength(opponent.ga_per_match),
                ]
            )
    rows.append(
        [
            "Media di lega",
            "",
            _fmt_strength(league.league_gf_per_match),
            _fmt_strength(league.league_ga_per_match),
        ]
    )
    _write_sheet(wb, "Calendario", headers, rows)


def _sheet_qualita_prezzo(wb: Workbook, remaining: Sequence[Player], league: LeagueContext) -> None:
    """Foglio "Qualità/prezzo": top 30 per punti per crediti spesi."""
    headers = ["Nome", "Squadra", "Ruolo", "QI", "Punti stagione", "Qualità/prezzo"]
    rows = sorted(
        (
            [
                player.name,
                player.team_name,
                ROLE_LABELS[player.role],
                player.quote.qi or 0,
                round(project(player, league).total_points, 1),
                _qp(player, league),
            ]
            for player in remaining
        ),
        key=lambda row: -row[5],
    )[:_TOP_QP]
    _write_sheet(wb, "Qualità prezzo", headers, rows)


def _player_row(player: Player, league: LeagueContext) -> list[object]:
    """Riga di proiezione per il foglio "Rimasti"."""
    proj = project(player, league)
    return [
        player.name,
        player.team_name,
        ROLE_LABELS[player.role],
        player.quote.qi or 0,
        round(proj.points_per_match, 2),
        round(proj.total_points, 1),
        _qp(player, league),
    ]


def _qp(player: Player, league: LeagueContext) -> float:
    """Qualità/prezzo: punti attesi per credito di quotazione."""
    qi = player.quote.qi or 0
    return round(project(player, league).total_points / qi, 2) if qi else 0.0


def _fmt_strength(value: float | None) -> str:
    return f"{value:.2f}" if value is not None else "—"


def _write_sheet(
    wb: Workbook,
    title: str,
    headers: list[str],
    rows: Sequence[Sequence[object]],
) -> None:
    """Scrive un foglio con intestazione in grassetto, freeze e autofit."""
    ws: Worksheet = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    for row in rows:
        ws.append(list(row))
    ws.freeze_panes = "A2"
    _autofit(ws)


def _autofit(ws: Worksheet, max_width: int = 40) -> None:
    """Larghezza colonne calcolata dal contenuto (con tetto)."""
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[letter].width = min(width, max_width)

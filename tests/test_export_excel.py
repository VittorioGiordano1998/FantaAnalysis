"""Test del report Excel (M6-T1): fogli, contenuti, salvataggio.

Nessun I/O reale: i bytes sono letti con openpyxl da BytesIO; il
salvataggio va su tmp_path.
"""

from io import BytesIO

from openpyxl import load_workbook

from entities import Player, Quote, Role, RoleGroup, SeasonStats
from export_excel import build_report, save_report
from optimize import optimize_squad
from projection import LeagueContext

LEAGUE = LeagueContext(season="2026-27", current_matchweek=1, teams={})

EXPECTED_SHEETS = [
    "Rosa ottimale",
    "Rimasti",
    "Classifica per ruolo",
    "Calendario",
    "Qualità prezzo",
]

SLOTS = {
    RoleGroup.P: 1,
    RoleGroup.D: 1,
    RoleGroup.C: 1,
    RoleGroup.A: 1,
}


def _player(url: str, role: Role, qi: int, points: float) -> Player:
    return Player(
        name=url,
        role=role,
        team_id="1",
        team_code="T",
        team_name="Team",
        quote=Quote(qi=qi, fvm=qi * 100),
        url=url,
        stats=SeasonStats(played_matches=10, fanta_avg=points / 38.0),
    )


PLAYERS = [
    _player("u1", Role.PC, 30, 70.0),
    _player("u2", Role.POR, 10, 50.0),
    _player("u3", Role.DC, 10, 55.0),
    _player("u4", Role.E, 10, 60.0),
    _player("u5", Role.A, 20, 65.0),
]


def _report_bytes(*, taken: frozenset[str] = frozenset()) -> bytes:
    squad = optimize_squad(PLAYERS, LEAGUE, slots=SLOTS, taken_urls=taken)
    return build_report(squad, PLAYERS, LEAGUE, taken)


def test_build_report_has_all_sheets():
    wb = load_workbook(BytesIO(_report_bytes()))
    assert wb.sheetnames == EXPECTED_SHEETS


def test_rosa_sheet_contains_selected_and_totals():
    wb = load_workbook(BytesIO(_report_bytes()))
    ws = wb["Rosa ottimale"]
    values = list(ws.values)
    assert len(values) == 6  # header + 4 scelti + TOTALE
    assert values[0][0] == "Nome"
    assert values[-1][0] == "TOTALE"
    assert values[-1][4] > 0


def test_rimasti_sheet_excludes_taken():
    wb = load_workbook(BytesIO(_report_bytes(taken=frozenset({"u1"}))))
    ws = wb["Rimasti"]
    names = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "u1" not in names
    assert len(names) == 4


def test_classifica_sheet_grouped_and_sorted():
    wb = load_workbook(BytesIO(_report_bytes()))
    ws = wb["Classifica per ruolo"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    groups = [row[0] for row in rows]
    assert groups == ["Attaccanti", "Attaccanti", "Centrocampisti", "Difensori", "Portieri"]
    assert [row[5] for row in rows] == [65, 70, 60, 55, 50]


def test_calendario_sheet_with_empty_league():
    squad = optimize_squad(PLAYERS, LEAGUE, slots=SLOTS)
    wb = load_workbook(BytesIO(build_report(squad, PLAYERS, LEAGUE, frozenset())))
    ws = wb["Calendario"]
    values = list(ws.values)
    assert values[0][0] == "Squadra"
    assert len(values) == 2  # header + media di lega


def test_qualita_prezzo_sheet_sorted_by_qp():
    wb = load_workbook(BytesIO(_report_bytes()))
    ws = wb["Qualità prezzo"]
    qp = [row[5] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert qp == sorted(qp, reverse=True)


def test_save_report_writes_file(tmp_path):
    data = _report_bytes()
    path = save_report(data, tmp_path / "report.xlsx")
    assert path.is_file()
    assert path.read_bytes() == data
